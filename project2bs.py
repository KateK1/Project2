from bs4 import BeautifulSoup
import openpyxl
import requests
import urllib.request
import re
import datetime

f = openpyxl.load_workbook('albumlist.xlsx')
sh = f['Worksheet']
sh['G1']= 'Country'  # Страна происхождения исполнителя
sh['H1']= 'StartY'  # Год начала карьеры
sh['I1']= 'FinY' # Год окончания карьеры

today = datetime.datetime.now()
t_year = str(today.year)



for row in range(2, sh.max_row):
    artist = str(sh.cell(row=row, column=4).value)
    print(artist)
    ar = artist.replace(' ', '_')
    req = requests.get('https://ru.wikipedia.org/wiki/'+ ar)
    if req.status_code != 404:
        req1 = urllib.request.urlopen('https://ru.wikipedia.org/wiki/' + ar)
        soup = BeautifulSoup(req1, features="lxml")
        if_death = False
        death_date = 0
        for table in soup.find_all('table', attrs = {"class": "infobox"}):
            for r in table.find_all('tr'):

                if re.search('Дата.смерти',r.text):
                    if_death = True
                    for d in r.find_all('td'):
                        death_date = d.text.strip().split('(')
                        death_date = str(death_date[0][-4:])
                        print(death_date)

                if r.text.find("Страна") >= 0:
                    for c in r.find_all('td'):
                        t = sh.cell(row=row, column=7)
                        country = str(c.text).strip().split('[')
                        print(str(country[0]))
                        t.value = str(country[0])

                if r.text.find('Год') >= 0:
                    for y in r.find_all('td'):
                        sy = sh.cell(row=row, column=8)
                        fy = sh.cell(row=row, column=9)
                        year = y.text.strip()
                        if 'н' in year:
                            year = str(year[0:5] + '-2021')
                        year = str(year).strip().split('[')
                        year = str(re.findall(r'\d+', year[0]))
                        sy.value = str(year)[2:6]
                        fy.value = str(year)[-6:-2]
                        if sy.value == fy.value:
                            if death_date:
                                fy.value = death_date
                            else:
                                fy.value = t_year
                        #print(str(year)[2:6])
                        #print(str(year)[-6:-2])
                        print(sy.value, fy.value)

                '''if r.text.find('Год') >= 0:
                    for y in r.find_all('td'):
                        sy = sh.cell(row=row, column=8)
                        fy = sh.cell(row=row, column=9)
                        year = y.text.strip()
                        if 'н' in year:
                            year = str(year[0:5]+'-'+t_year)
                        if len(year) < 7:
                            if if_death:
                                fy.value = death_date
                                if_death = False
                            else:
                                fy.value = t_year
                                if_death = False'''





f.save('albumlist.xlsx')