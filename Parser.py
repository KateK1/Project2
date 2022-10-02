from bs4 import BeautifulSoup
import openpyxl
import requests
import urllib.request
import re

f = openpyxl.load_workbook('albumlist.xlsx')
sh = f['Worksheet']
sh['G1']= 'Country'
sh['H1']= 'StartY'
sh['I1']= 'FinY'


for row in range(2, sh.max_row):
    artist = (str(sh.cell(row=row, column=4).value)).replace(' ', '_')
    print(artist)
    req = requests.get('https://ru.wikipedia.org/wiki/'+ artist)
    if req.status_code != 404:
        req1 = urllib.request.urlopen('https://ru.wikipedia.org/wiki/' + artist)
        soup = BeautifulSoup(req1, features="lxml")
        for table in soup.find_all('table', attrs = {"class": "infobox"}):
            for r in table.find_all('tr'):
                # find country for each artist
                if r.text.find("Страна") >= 0:
                    for c in r.find_all('td'):
                        t = sh.cell(row=row, column=7)
                        country = str(c.text).strip().split('[')
                        print(str(country[0]))
                        t.value = str(country[0])
                # find year when artist started career
                if r.text.find('Год') >= 0:
                    for y in r.find_all('td'):
                        t = sh.cell(row=row, column=8)
                        b = sh.cell(row=row, column=9)
                        year = y.text.strip()
                        # for artists who still active 'н' stands for 'Until now'
                        if 'н' in year:
                            year = str(year[0:5]+'-2021')
                        year = str(year).strip().split('[')
                        year = str(re.findall(r'\d+',year[0]))
                        t.value = str(year)[2:6]
                        b.value = str(year)[-6:-2]
                        print(str(year)[2:6])
                        print(str(year)[-6:-2])

f.save('albumlist.xlsx')
